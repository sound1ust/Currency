import os
from datetime import datetime
from logging import getLogger

import openpyxl
from celery import shared_task
from converter.models import Converter

from .models import UserTask
from .models import UserTaskResult
from .utils import format_result_name
from .utils import format_start_parameters
from currency import settings

logger = getLogger('currency_user_tasks_manager')


@shared_task
def export_task(task_id):
    time_before = datetime.now()
    task = UserTask.objects.filter(id=task_id).first()
    if not task:
        logger.warning(f'Warning: {__name__}: No one task in chosen')
        return None

    # task.update_task_status('Started')
    start_parameters = {
        'created_at__gte': task.date_from,
        'created_at__lte': task.date_to,
        'input_ticker__in': task.input_tickers,
        'output_ticker__in': task.output_tickers,
        'value__gte': task.value_from,
        'value__lte': task.value_to,
        'source_id__in': task.sources.all(),
        'updated_by__in': task.updated_by.all(),
    }
    start_parameters = {
        key: value for key, value in start_parameters.items() if value
    }

    task_result = UserTaskResult.objects.create(
        name=format_result_name(task.name, time_before),
        user=task.user,
        start_date=time_before,
        start_parameters=format_start_parameters(start_parameters),
    )

    try:
        converters = Converter.objects.filter(**start_parameters)
        if not converters:
            task_result.status = 'FAILURE'
            task_result.save()

        task_result.save()
        create_excel(task_result, converters)

    except Exception as exc:
        logger.error(f'Error: {__name__}: {exc}')
        task_result.status = 'FAILURE'
        task_result.save()


def create_excel(task_result, converters):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = f'Currencies of {task_result.name}'

    header = converters.first().get_field_names()
    for col_num, column_title in enumerate(header, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = str(column_title)

    for row_num, row in enumerate(converters.values_list(), 1):
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num + 1, column=col_num)
            cell.value = str(cell_value)

    file_name = f'{task_result.name}.xlsx'
    file_dir = os.path.join(settings.BASE_DIR, 'excel')
    if not os.path.exists(file_dir):
        os.mkdir(file_dir)

    file_path = os.path.join(file_dir, file_name)

    workbook.save(file_path)

    task_result.result_link = file_path
    task_result.status = 'SUCCEED'
    time_after = datetime.now()
    task_result.duration = time_after - task_result.start_date
    task_result.save()
